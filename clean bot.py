import telebot
from datetime import timedelta, datetime
from telebot import types
import openpyxl
import pandas as pd


#пользователи
class Users:
    def __init__(self, flag):
        self.flag = flag
        self.data = []
        self.flag_data = False
        self.week = ''
        self.day = ''
        self.teacher_letter = ''
        self.teacher = ''
        self.class_num = ''
        self.class_num_letter = ''
        self.cabinet = ''
        self.floor = ''


    def reflag(self, flag):
        self.flag = flag

    def choose_week(self, a):
        self.week = a

    def choose_day(self, a):
        self.day = a

    def choose_teacher_letter(self, a):
        self.teacher_letter = a

    def choose_teacher(self, a):
        self.teacher = a

    def choose_class_num(self, a):
        self.class_num = a

    def choose_class_num_letter(self, a):
        self.class_num_letter = a

    def choose_cabinet(self, a):
        self.cabinet = a

    def choose_floor(self, a):
        self.floor = a

    def reset_flag_data(self):
        if self.flag_data == False:
            self.flag_data = True
        else:
            self.flag_data = False


#функции
def week_choose(message):
    keyboard_week = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_even = types.KeyboardButton('Четная')
    key_uneven = types.KeyboardButton('Нечетная')
    key_tomorrow = types.KeyboardButton('На завтра')
    key_even_week = types.KeyboardButton('На всю четную неделю')
    key_uneven_week = types.KeyboardButton('На всю нечетную неделю')
    key_back = types.KeyboardButton('В начало')
    keyboard_week.add(key_even, key_uneven, key_even_week, key_uneven_week, key_tomorrow, key_back)
    bot.send_message(message.chat.id, 'Выберите неделю, либо получите расписание на завтра', reply_markup=keyboard_week)
    return

def day_choose(message):
    keyboard_day = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_monday = types.KeyboardButton('Понедельник')
    key_tuesday = types.KeyboardButton('Вторник')
    key_wednesday = types.KeyboardButton('Среда')
    key_thursday = types.KeyboardButton('Четверг')
    key_friday = types.KeyboardButton('Пятница')
    key_saturday = types.KeyboardButton('Суббота')
    key_back = types.KeyboardButton('В начало')
    keyboard_day.add(key_monday, key_tuesday, key_wednesday, key_thursday, key_friday, key_saturday, key_back)
    bot.send_message(message.chat.id, "Выберите день недели", reply_markup=keyboard_day)
    return

def day_check(a):
    if a == 'Понедельник' or a == 'Вторник' or a == 'Среда' or a == 'Четверг' or a == 'Пятница' or a == 'Суббота':
        return True
    else:
        return False

def class_num_choose(message):
    keyboard_num = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_8 = types.KeyboardButton('8')
    key_9 = types.KeyboardButton('9')
    key_10 = types.KeyboardButton('10')
    key_11 = types.KeyboardButton('11')
    key_back = types.KeyboardButton('В начало')
    keyboard_num.add(key_8, key_9, key_10, key_11, key_back)
    bot.send_message(message.chat.id, "Выберите нужный вам класс", reply_markup=keyboard_num)
    return

def class_num_check(a):
    if a == '8' or a == '9' or a == '10' or a == '11':
        return True
    else:
        return False

def class_letter_choose(message, num):
    keyboard_letter = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_m1 = types.KeyboardButton('М1')
    key_m2 = types.KeyboardButton('М2')
    key_n1 = types.KeyboardButton('Н1')
    key_n2 = types.KeyboardButton('Н2')
    key_o1 = types.KeyboardButton('О1')
    key_o2 = types.KeyboardButton('О2')
    key_p1 = types.KeyboardButton('П1')
    key_p2 = types.KeyboardButton('П2')
    key_r1 = types.KeyboardButton('Р1')
    key_r2 = types.KeyboardButton('Р2')
    key_s = types.KeyboardButton('С')
    key_back = types.KeyboardButton('В начало')
    if num == '9' or num == '11':
        keyboard_letter.add(key_m1, key_m2, key_n1, key_n2, key_o1, key_o2, key_p1, key_p2, key_r1, key_r2)
    elif num == '8':
        keyboard_letter.add(key_m1, key_m2, key_n1, key_n2, key_o1, key_o2)
    elif num == '10':
        keyboard_letter.add(key_m1, key_m2, key_n1, key_n2, key_o1, key_o2, key_p1, key_p2, key_r1, key_r2, key_s)
    keyboard_letter.add(key_back)
    bot.send_message(message.chat.id, "Выберите свой класс и группу", reply_markup=keyboard_letter)
    return

def floor_choose(message):
    keyboard_floor = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_1 = types.KeyboardButton('1')
    key_2 = types.KeyboardButton('2')
    key_3 = types.KeyboardButton('3')
    key_back = types.KeyboardButton('В начало')
    keyboard_floor.add(key_1, key_2, key_3, key_back)
    bot.send_message(message.chat.id, "Выберите этаж", reply_markup=keyboard_floor)

def floor_check(a):
    if a == '1' or a == '2' or a == '3':
        return True
    else:
        return False

def cab_num_choose(message, floor):
    keyboard_cab_num = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if floor == '1':
        cabs = open('1 floor.txt', 'r')
    elif floor == '2':
        cabs = open('2 floor.txt', 'r')
    elif floor == '3':
        cabs = open('3 floor.txt', 'r')
    a = ''
    for cab in cabs:
        a += cab
    a = list(a.split(' '))
    key_list = [types.KeyboardButton(i) for i in a]
    key_back = types.KeyboardButton('В начало')
    keyboard_cab_num.add(*key_list, key_back)
    bot.send_message(message.chat.id, 'Выберите кабинет', reply_markup=keyboard_cab_num)

def teacher_letter_choose(message):
    a = ['А', 'Б', 'В', 'Г', 'Д', 'З', 'И', 'К', 'Л', 'М', 'Н', 'Р', 'С', 'У', 'Ф', 'Х', 'Ч', 'Ш', 'Я']
    keyboard_teacher_letter = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_list = [types.KeyboardButton(i) for i in a]
    key_back = types.KeyboardButton('В начало')
    keyboard_teacher_letter.add(*key_list, key_back)
    bot.send_message(message.chat.id, 'Выберите первую букву фамилии учителя', reply_markup=keyboard_teacher_letter)

def teacher_letter_check(message):
    a = ['А', 'Б', 'В', 'Г', 'Д', 'З', 'И', 'К', 'Л', 'М', 'Н', 'Р', 'С', 'У', 'Ф', 'Х', 'Ч', 'Ш', 'Я']
    if message.text in a:
        return True
    else:
        return False

def teacher_choose(message, letter):
    keyboard_teacher = types.ReplyKeyboardMarkup(resize_keyboard=True)
    file = letter+'.txt'
    f = open(file, encoding="utf-8")
    for line in f:
        key_list = [types.KeyboardButton(i) for i in line.split(', ')]
    key_back = types.KeyboardButton('В начало')
    keyboard_teacher.add(*key_list, key_back)
    bot.send_message(message.chat.id, 'Выберите учителя', reply_markup=keyboard_teacher)

def tomorrow_set(message):
    global users
    #день
    day_num = datetime.today().weekday()
    days = {
        '0': 'Понедельник',
        '1': 'Вторник',
        '2': 'Среда',
        '3': 'Четверг',
        '4': 'Пятница',
        '5': 'Суббота',
        '6': 'Воскресенье',
    }
    if day_num+1 == 6:
        bot.send_message(message.chat.id, 'Завтра воскресенье, будет показано расписание на понедельник', reply_markup=None)
        users[message.chat.id].choose_day(days['0'])
    else:
        users[message.chat.id].choose_day(days[str(day_num+1)])
    #неделя
    a = int(str(datetime.now()).split()[0].split('-')[-1])
    year1 = timedelta(days=a)
    year2 = timedelta(days=10)

    difference = year1 - year2
    b = difference.days

    if (b // 7) % 2 == 0:
        users[message.chat.id].choose_week('Четная')
    else:
        users[message.chat.id].choose_week('Нечетная')

def last_item(message):
    global users

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    key_yes = types.KeyboardButton('Да')
    key_back = types.KeyboardButton('В начало')
    keyboard.add(key_yes, key_back)

    users[message.chat.id].data.append(message.text)
    users[message.chat.id].reflag('data')
    users[message.chat.id].reset_flag_data()
    bot.send_message(message.chat.id, 'Данные обновлены, посмотреть расписание на завтра?', reply_markup=keyboard)

def cabinet(week, cab, day):
    header = [['8М1', '8М2', '8Н1', '8Н2', '8О1', '8О2'],
              ['10М1', '10М2', '10Н1', '10Н2', '10О1', '10О2', '10П1', '10П2', '10Р1', '10Р2', '10C'],
              ['9М1', '9М2', '9Н1', '9Н2', '9О1', '9О2', '9П1', '9П2', '9Р1', '9Р2'],
              ['11М1', '11М2', '11Н1', '11Н2', '11О1', '11О2', '11П1', '11П2', '11Р1']]

    # создание dataframe на нужную неделю
    data_list = []
    wb = openpyxl.load_workbook("rasp_s_10_yanvarya_1.xlsx")
    for sheet_name in wb.sheetnames:
        if sheet_name.split()[-1] == week[:-3].lower():
            df = pd.read_excel("rasp_s_10_yanvarya_1.xlsx", sheet_name=sheet_name)
            data_list.append(df)


    # создание списка с данными на конкретный день
    day_data = []
    for df in data_list:
        data = df.values.tolist()
        for string in data:
            if string[0] == day.lower():
                day_data.append(string)


    # поиск кабинета
    locations = []
    x = 0
    y = 0
    for string in day_data:
        for item in string:
            if isinstance(item, str):
                if item.split()[-1] == cab:
                    locations.append([x, y])
            x += 1
        y += 1
        x = 0


    output = []
    for i in locations:
        if day.lower() == "суббота":
            output.append(day_data[i[1]][1][0:7] + day_data[i[1] + 1][1][7:] + ' ' + day_data[i[1]][i[0]] + ' ' + day_data[i[1]+1][i[0]] + ' ' + header[i[1] // 7][i[0] - 2])
        else:
            output.append(day_data[i[1]][1][0:7] + day_data[i[1] + 1][1][7:] + ' ' + day_data[i[1]][i[0]] + ' ' + day_data[i[1]+1][i[0]] + ' ' + header[i[1] // 9][i[0] - 2])

    # вывод
    i = 0
    while i < len(output)-1:
        if output[i][:-1] == output[i+1][:-1]:
            output.pop(i+1)
            output[i] = output[i][:-1]
        i += 1
    output.sort()
    output_s = day.upper()
    if output == []:
        output_s += '\n\nРасписание на этот день отсутствует'
    else:
        for i in output:
            output_s += '\n' + i
    return output_s

def clas_s(week, class_num, day):
    # создание dataframe для нужного класса на нужную неделю
    wb = openpyxl.load_workbook("rasp_s_10_yanvarya_1.xlsx")
    for sheet_name in wb.sheetnames:
        if sheet_name.split()[-1] == week[:-3].lower() and sheet_name[0] == class_num[0]:
            df = pd.read_excel("rasp_s_10_yanvarya_1.xlsx", sheet_name=sheet_name)
            break

    # создание списка с данными на конкретный день
    data = df[["день", "время", class_num.upper()]].values.tolist()[1:]
    day_data = []
    for string in data:
        if string[0] == day.lower():
            day_data.append(string)

    # вывод данных
    output = day_data[0][0].upper()
    for string in day_data:
        if isinstance(string[2], str):
            output += '\n' + string[1] + ': ' + string[2]

    if output == day.upper():
        output += '\n\nРасписание на этот день отсутствует'
    return output

def teacher(week, teacher, day):
    data_list = []
    header = [['8М1', '8М2', '8Н1', '8Н2', '8О1', '8О2'],
              ['10М1', '10М2', '10Н1', '10Н2', '10О1', '10О2', '10П1', '10П2', '10Р1', '10Р2', '10C'],
              ['9М1', '9М2', '9Н1', '9Н2', '9О1', '9О2', '9П1', '9П2', '9Р1', '9Р2'],
              ['11М1', '11М2', '11Н1', '11Н2', '11О1', '11О2', '11П1', '11П2', '11Р1']]

    # создание dataframe на нужную неделю
    wb = openpyxl.load_workbook("rasp_s_10_yanvarya_1.xlsx")
    for sheet_name in wb.sheetnames:
        if sheet_name.split()[-1] == week[:-3].lower():
            df = pd.read_excel("rasp_s_10_yanvarya_1.xlsx", sheet_name=sheet_name)
            data_list.append(df)

    # создание списка с данными на конкретный день
    day_data = []
    for df in data_list:
        data = df.values.tolist()
        for string in data:
            if string[0] == day.lower():
                day_data.append(string)

    # поиск учителя
    locations = []
    x = 0
    y = 0
    for string in day_data:
        for item in string:
            if isinstance(item, str):
                # print(item, x, y)
                if item == teacher:
                    locations.append([x, y])
                    # print(item, x, y)
            x += 1
        y += 1
        x = 0

    output = []
    for i in locations:
        if day.lower() == 'суббота':
            output.append(day_data[i[1] - 1][1][0:7] + day_data[i[1]][1][7:] + ' ' + day_data[i[1] - 1][i[0]] + ' ' +
                          header[i[1] // 7][i[0] - 2])
        else:
            output.append(day_data[i[1] - 1][1][0:7] + day_data[i[1]][1][7:] + ' ' + day_data[i[1] - 1][i[0]] + ' ' +
                          header[i[1] // 9][i[0] - 2])
    i = 0
    while i < len(output) - 1:
        if output[i][:-1] == output[i + 1][:-1]:
            output.pop(i + 1)
            output[i] = output[i][:-1]
        i += 1

    output.sort()
    output_s = day.upper()
    if output == []:
        output_s += '\n\nРасписание на этот день отсутствует'
    else:
        for i in output:
            output_s += '\n' + i
    return output_s






#главная часть
bot=telebot.TeleBot('2058610379:AAFajD6GA1NrzXfhZYPm9An_THZOvkx7YdA')
flag = ''
users = {}

keyboard_category = types.ReplyKeyboardMarkup(resize_keyboard=True)
keyboard_category_data_choose = types.ReplyKeyboardMarkup(resize_keyboard=True)
key_teacher = types.KeyboardButton('На учителя')
key_class_day = types.KeyboardButton('На класс')
key_cabinet = types.KeyboardButton('На кабинет')
key_data = types.KeyboardButton('Свои данные')
keyboard_category.add(key_teacher, key_class_day, key_cabinet, key_data)
keyboard_category_data_choose.add(key_teacher, key_class_day, key_cabinet)

@bot.message_handler(commands=['start', 'help'])
def category(message):
    global keyboard_category
    bot.reply_to(message, "Привет! Какое нужно расписание?", reply_markup=keyboard_category)
    users[message.chat.id] = Users('choose')


@bot.message_handler(content_types=['text'])
def subcategory(message):
    global keyboard_category
    global users
    try:
        if message.text == "На учителя" and users[message.chat.id].flag == 'choose':
            teacher_letter_choose(message)
            users[message.chat.id].reflag('teacher_letter')
            if users[message.chat.id].flag_data == True:
                users[message.chat.id].data.append('teacher')

        elif message.text == "На класс" and users[message.chat.id].flag == 'choose':
            class_num_choose(message)
            users[message.chat.id].reflag('class_num')
            if users[message.chat.id].flag_data == True:
                users[message.chat.id].data.append('class')

        elif message.text == "На кабинет" and users[message.chat.id].flag == 'choose':
            floor_choose(message)
            users[message.chat.id].reflag('cabinet_floor')
            if users[message.chat.id].flag_data == True:
                users[message.chat.id].data.append('cab')

        elif message.text == "Свои данные" and users[message.chat.id].flag == 'choose':
            users[message.chat.id].reflag('data')
            if users[message.chat.id].data != []:
                keyboard_data = types.ReplyKeyboardMarkup(resize_keyboard=True)
                key_receive = types.KeyboardButton('Получить расписание')
                key_rechoose = types.KeyboardButton('Изменить параметры')
                key_back = types.KeyboardButton('В начало')
                keyboard_data.add(key_receive, key_rechoose, key_back)
                bot.send_message(message.chat.id,
                                 'Получите расписание либо измените параметры, по которым вы хотите получать расписание',
                                 reply_markup=keyboard_data)
            if users[message.chat.id].data == []:
                users[message.chat.id].reset_flag_data()
                bot.send_message(message.chat.id, 'Укажите параметры, по которым вы хотите получать расписание')
                bot.send_message(message.chat.id, "Какое нужно расписание?", reply_markup=keyboard_category_data_choose)
                users[message.chat.id].reflag('choose')

        elif users[message.chat.id].flag == 'data' and (message.text == 'Да' or message.text == 'Получить расписание'):
            if users[message.chat.id].data[0] == "teacher":
                users[message.chat.id].choose_teacher(users[message.chat.id].data[1])
                tomorrow_set(message)
                bot.send_message(message.chat.id, teacher(users[message.chat.id].week, users[message.chat.id].teacher, users[message.chat.id].day,))
            elif users[message.chat.id].data[0] == "class":
                users[message.chat.id].choose_class_num(users[message.chat.id].data[1])
                users[message.chat.id].choose_class_num_letter(users[message.chat.id].data[2])
                tomorrow_set(message)
                bot.send_message(message.chat.id, clas_s(users[message.chat.id].week,
                                                         users[message.chat.id].class_num_letter, users[message.chat.id].day))
            elif users[message.chat.id].data[0] == "cab":
                users[message.chat.id].choose_cabinet(users[message.chat.id].data[1])
                tomorrow_set(message)
                bot.send_message(message.chat.id,
                                 cabinet(users[message.chat.id].week, users[message.chat.id].cabinet, users[message.chat.id].day))
            bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
            users[message.chat.id].reflag('choose')

        elif message.text == 'Изменить параметры':
            users[message.chat.id].data = []
            users[message.chat.id].reset_flag_data()
            bot.send_message(message.chat.id, 'Укажите параметры, по которым вы хотите получать расписание')
            bot.send_message(message.chat.id, "Какое нужно расписание?", reply_markup=keyboard_category_data_choose)
            users[message.chat.id].reflag('choose')

        elif message.text == 'В начало':
            bot.send_message(message.chat.id, "Привет! Какое нужно расписание?", reply_markup=keyboard_category)
            users[message.chat.id].reflag('choose')
            if users[message.chat.id].flag_data == True:
                users[message.chat.id].flag_data = False
                users[message.chat.id].data = []

    # ввод и вывод данных
    #учитель
        elif users[message.chat.id].flag == 'teacher_letter':
            if teacher_letter_check(message):
                users[message.chat.id].choose_teacher_letter(message.text)
                teacher_choose(message, users[message.chat.id].teacher_letter)
                users[message.chat.id].reflag('teacher')

        elif users[message.chat.id].flag == 'teacher':
            users[message.chat.id].choose_teacher(message.text)
            users[message.chat.id].reflag('teacher_week')
            if users[message.chat.id].flag_data == True:
                last_item(message)
            else:
                week_choose(message)

        elif users[message.chat.id].flag == 'teacher_week':
            if message.text == 'Четная':
                users[message.chat.id].choose_week('Четная')
                users[message.chat.id].reflag('teacher_day')
                day_choose(message)
            elif message.text == 'Нечетная':
                users[message.chat.id].choose_week('Нечетная')
                users[message.chat.id].reflag('teacher_day')
                day_choose(message)
            elif message.text == 'На всю четную неделю':
                users[message.chat.id].choose_week('Четная')
                for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                    bot.send_message(message.chat.id,
                                     teacher(users[message.chat.id].week, users[message.chat.id].teacher, day),
                                     reply_markup=None)
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            elif message.text == 'На всю нечетную неделю':
                users[message.chat.id].choose_week('Нечетная')
                for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                    bot.send_message(message.chat.id,
                                     teacher(users[message.chat.id].week, users[message.chat.id].teacher, day),
                                     reply_markup=None)
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            elif message.text == 'На завтра':
                tomorrow_set(message)
                bot.send_message(message.chat.id, teacher(users[message.chat.id].week, users[message.chat.id].teacher, users[message.chat.id].day))
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            else:
                week_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")

        elif users[message.chat.id].flag == 'teacher_day':
            if day_check(message.text):
                users[message.chat.id].choose_day(message.text)
                bot.send_message(message.chat.id, teacher(users[message.chat.id].week, users[message.chat.id].teacher, users[message.chat.id].day))
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            else:
                day_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")


    #класс
        elif users[message.chat.id].flag == 'class_num':
            if class_num_check(message.text):
                users[message.chat.id].choose_class_num(message.text)
                users[message.chat.id].reflag('class_letter')
                class_letter_choose(message, users[message.chat.id].class_num)
                if users[message.chat.id].flag_data == True:
                    users[message.chat.id].data.append(message.text)
            else:
                class_num_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")

        elif users[message.chat.id].flag == 'class_letter':
            users[message.chat.id].choose_class_num_letter(users[message.chat.id].class_num + message.text)
            users[message.chat.id].reflag('class_week')
            if users[message.chat.id].flag_data == True:
                last_item(message)
                users[message.chat.id].data[2] = users[message.chat.id].data[1] + users[message.chat.id].data[2]
            else:
                week_choose(message)

        elif users[message.chat.id].flag == 'class_week':
            if message.text == 'Четная':
                users[message.chat.id].choose_week('Четная')
                users[message.chat.id].reflag('class_day')
                day_choose(message)
            elif message.text == 'Нечетная':
                users[message.chat.id].choose_week('Нечетная')
                users[message.chat.id].reflag('class_day')
                day_choose(message)
            elif message.text == 'На всю четную неделю':
                users[message.chat.id].choose_week('Четная')
                for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                    bot.send_message(message.chat.id, clas_s(users[message.chat.id].week,
                                                             users[message.chat.id].class_num_letter, day), reply_markup=None)
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            elif message.text == 'На всю нечетную неделю':
                users[message.chat.id].choose_week('Нечетная')
                for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                    bot.send_message(message.chat.id, clas_s(users[message.chat.id].week,
                                                             users[message.chat.id].class_num_letter, day), reply_markup=None)
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            elif message.text == 'На завтра':
                tomorrow_set(message)
                bot.send_message(message.chat.id, clas_s(users[message.chat.id].week, users[message.chat.id].class_num_letter, users[message.chat.id].day))
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            else:
                week_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")

        elif users[message.chat.id].flag == 'class_day':
            if day_check(message.text):
                users[message.chat.id].choose_day(message.text)
                users[message.chat.id].reflag('choose')
                bot.send_message(message.chat.id, clas_s(users[message.chat.id].week, users[message.chat.id].class_num_letter, users[message.chat.id].day))
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
            else:
                day_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")


    #кабинет
        elif users[message.chat.id].flag == 'cabinet_floor':
            if floor_check(message.text):
                users[message.chat.id].choose_floor(message.text)
                users[message.chat.id].reflag('cabinet_num')
                cab_num_choose(message, users[message.chat.id].floor)
            else:
                floor_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")

        elif users[message.chat.id].flag == 'cabinet_num':
            users[message.chat.id].choose_cabinet(message.text)
            users[message.chat.id].reflag('cabinet_week')
            if users[message.chat.id].flag_data == True:
                last_item(message)
            else:
                week_choose(message)

        elif users[message.chat.id].flag == 'cabinet_week':
            if message.text == 'Четная':
                users[message.chat.id].choose_week('Четная')
                users[message.chat.id].reflag('cabinet_day')
                day_choose(message)
            elif message.text == 'Нечетная':
                users[message.chat.id].week = ('Нечетная')
                users[message.chat.id].reflag('cabinet_day')
                day_choose(message)
            elif message.text == 'На всю четную неделю':
                users[message.chat.id].choose_week('Четная')
                for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                    bot.send_message(message.chat.id, cabinet(users[message.chat.id].week, users[message.chat.id].cabinet, day), reply_markup=None)
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            elif message.text == 'На всю нечетную неделю':
                users[message.chat.id].choose_week('Нечетная')
                for day in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                    bot.send_message(message.chat.id, cabinet(users[message.chat.id].week, users[message.chat.id].cabinet, day), reply_markup=None)
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            elif message.text == 'На завтра':
                tomorrow_set(message)
                bot.send_message(message.chat.id,
                                 cabinet(users[message.chat.id].week, users[message.chat.id].cabinet, users[message.chat.id].day))
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            else:
                week_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")

        elif users[message.chat.id].flag == 'cabinet_day':
            if day_check(message.text):
                users[message.chat.id].choose_day(message.text)
                bot.send_message(message.chat.id,
                                 cabinet(users[message.chat.id].week, users[message.chat.id].cabinet, users[message.chat.id].day))
                bot.send_message(message.chat.id, 'Какое расписание еще показать?', reply_markup=keyboard_category)
                users[message.chat.id].reflag('choose')
            else:
                day_choose(message)
                bot.send_message(message.chat.id, "Ошибка! Введите повторно!")


    except Exception:
        users[message.chat.id] = Users('choose')
        print('exeption')


bot.polling()